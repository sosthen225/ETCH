from functools import partial
from urllib import request
from django.forms import ValidationError, formset_factory, inlineformset_factory, modelformset_factory
from django.shortcuts import get_object_or_404, render
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib import messages
from Geequipe.forms import ActiviteForm, AffectationProjetForm, EquipeForm, LivrableForm, LivrableInlineFormSet, MembreFormSet, MobilisationForm, ModifierPersonnelForm, ProjetForm
from Geequipe.models import ChefProjet, Projet , Client
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from datetime import date, datetime, timedelta, timezone
from django.shortcuts import render, redirect
from django.contrib.auth import logout, authenticate, login
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_protect
from Geequipe.models import ChefProjet, Projet, Client
import json
from datetime import datetime
from django.http import JsonResponse
from .models import COMPETENCE_CHOICES, PAYS_CHOICES, STATUT_ACTIVITIES_CHOICES, STATUT_PERSONNEL_CHOICES, Activite, AffectationProjet, Certificat, Competence, Equipe, Expatriation, Livrable, Membre, Mobilisation, PaysAffectation, Personnel, Projet, Client, Posseder, Realiser
from django.contrib.auth import get_user_model
User = get_user_model()
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Projet, Client, ChefProjet
from django.contrib.auth.models import User
from django.views.decorators.http import require_POST 
from django.utils.dateparse import parse_date
from django.core.files.uploadedfile import UploadedFile
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.utils.dateparse import parse_date
from django.contrib import messages





@csrf_exempt
def login_page(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            # Vérifie si cet utilisateur est un chef de projet
            if ChefProjet.objects.filter(user=user).exists():
                login(request, user)
                messages.success(request, "Connexion réussie en tant que Chef de Projet !")
                return redirect('index')  # Redirige vers la page d'accueil
            else:
                messages.error(request, "Accès réservé aux Chefs de Projet.")
        else:
            messages.error(request, "Identifiants incorrects.")

    return render(request, 'login.html')






def index(request):

     # --- Mettre à jour le statut des projets avant de les récupérer ---
    # Récupérer tous les projets en cours
    projets_en_cours_a_verifier = Projet.objects.filter(statut='en_cours')

    # Parcourir ces projets et mettre à jour ceux qui sont terminés
    # Il est plus efficace de faire une mise à jour en masse si possible, mais une boucle est plus claire pour la démonstration.
    updated_count = 0
    today = timezone.localdate() # Utiliser timezone.localdate() pour la date locale

    for projet in projets_en_cours_a_verifier:
        if projet.date_fin <= today:
            projet.statut = 'terminé'
            projet.save()
            updated_count += 1

    if updated_count > 0:
        print(f"{updated_count} projets mis à jour vers le statut 'terminé'.")


    total_teams = Equipe.objects.count()
    total_personnel = Personnel.objects.count()
    completed_projects = Projet.objects.filter(statut='terminé').count()
    ongoing_projects = Projet.objects.filter(statut='en_cours').count()
    projet_encours= Projet.objects.filter(statut='en_cours').order_by('-date_fin')[:10]
    completed_projects_list = Projet.objects.filter(statut='terminé').order_by('-date_fin')[:10] # Obtenir les 10 derniers projets terminés

    # Pour les notifications : Projets se terminant bientôt (par exemple, dans une semaine)
    today = date.today()
    one_month_from_now = today + timedelta(days=7)
    upcoming_projects = Projet.objects.filter(
    date_fin__gt=today,
    date_fin__lte=one_month_from_now,
    statut='en_cours'
).order_by('date_fin')
    
    #all_mobilisations = Mobilisation.objects.all().order_by('-date_debut')
    #planned_mobilisations = Mobilisation.objects.filter(statut__in=['planifié'] ).order_by('date_debut')
    planned_activities =Activite.objects.filter(statut='planifiée' ,projet__statut='en_cours').order_by('-date_debut')


    # # Pour les notifications : Certifications du personnel expirant bientôt (par exemple, dans les 30 jours)
    # expiring_personnel = Personnel.objects.filter(
    #     certification_expiry_date__gt=today,
    #     certification_expiry_date__lte=one_month_from_now
    # ).order_by('certification_expiry_date')


    context = {
        'total_teams': total_teams,
        'ongoing_projects': ongoing_projects,
        'completed_projects': completed_projects,
        'total_personnel': total_personnel,
        'completed_projects_list': completed_projects_list,
        'upcoming_projects': upcoming_projects,
        'projet_encours': projet_encours,
        'planned_activities': planned_activities
        #'planned_mobilisations':planned_mobilisations
       # 'expiring_personnel': expiring_personnel,
    }
    return render(request, 'index.html',context)







def tabpersonels(request):
    agents = Personnel.objects.all()
    return render(request, 'data_personels.html',{ 'competences': COMPETENCE_CHOICES,
        'agents': agents,
        'pays_liste':  PAYS_CHOICES,
        'STATUT_PERSONNEL_CHOICES': STATUT_PERSONNEL_CHOICES,
        'certificats': Certificat.objects.all(),})



def tabprojet(requests):
    projets = Projet.objects.all()
    chefs_de_projet = User.objects.all()
    clients = Client.objects.all()
    return render(requests, 'projet.html', {
        'projets': projets,
        'chefs_de_projet': chefs_de_projet,
        'clients': clients,
       
    })

def modifier_statut_projet(request, projet_id):
    if request.method == 'POST':
        nouveau_statut = request.POST.get('statut')
        try:
            projet = Projet.objects.get(id=projet_id)
            projet.statut = nouveau_statut
            projet.save()
            return JsonResponse({'success': True})
        except Projet.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Projet introuvable'})
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})


def changer_statut_personnel(request, agent_id):
    if request.method == 'POST':
        agent = get_object_or_404(Personnel, id=agent_id)
        nouveau_statut = request.POST.get('statut')
        if nouveau_statut in dict(STATUT_PERSONNEL_CHOICES):
            agent.statut = nouveau_statut
            agent.save()
            messages.success(request, "Statut de l'agent mis à jour.")
        else:
            messages.error(request, 'Statut non valide.')
    return redirect('tabpersonels') 





User = get_user_model()
@csrf_exempt
def ajouter_projet(request):
    if request.method == 'POST':
        try:
            print("Début ajout projet")
            data = json.loads(request.body)
            print("Données reçues:", data)

            client_nom = data.get('client_nom', '').strip()
            client, _ = Client.objects.get_or_create(nom=client_nom)
            print(f"Client récupéré/créé : {client}")

            user = User.objects.get(id=int(data.get('chef_projet')))
            chef_projet = ChefProjet.objects.get(user=user)
            print(f"Chef de projet récupéré : {chef_projet}")

            projet = Projet(
                nom=data.get('nom'),
                type=data.get('type'),
                date_debut=data.get('date_debut'),
                date_fin=data.get('date_fin'),
                site=data.get('site'),
                ville=data.get('ville'),
                pays=data.get('pays'),
                statut=data.get('statut'),
                client=client,
                chef_projet=chef_projet
            )

            # Validation personnalisée
            projet.clean()

            # Sauvegarde après validation
            projet.save()
            print("Projet créé avec succès")

            return JsonResponse({'success': True})

        except ValidationError as ve:
            print("Erreur de validation :", ve.message)
            return JsonResponse({'success': False, 'error': ve.message}, status=400)
            
        except User.DoesNotExist:
            print("Erreur : Utilisateur introuvable")
            return JsonResponse({'success': False, 'error': 'Utilisateur introuvable'}, status=400)
        except ChefProjet.DoesNotExist:
            print("Erreur : Chef de projet introuvable")
            return JsonResponse({'success': False, 'error': 'Chef de projet introuvable'}, status=400)
        except Exception as e:
            print("Erreur dans ajouter_projet:", e)
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    print("Erreur : Méthode non autorisée")
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)




@csrf_exempt

def modifier_projet(request,id):
    projet = get_object_or_404(Projet, id=id)
    if request.method == 'POST':
        form = ProjetForm(request.POST, instance=projet)
        if form.is_valid():
            form.save()
            return redirect('tabprojet')
    else:
        form = ProjetForm(instance=projet)
    return render(request, 'modifier_projet.html', {'form': form})




def supprimer_projet(request, id):
    projet = get_object_or_404(Projet, id=id)
    if request.method == 'POST':
        projet.delete()
    return redirect('tabprojet')







from django.forms.models import model_to_dict

def projet_json(request, projet_id):
    projet = get_object_or_404(Projet, id=projet_id)
    data = model_to_dict(projet)
    # exemple retourner l'id du chef de projet et le nom client séparément
    data['chef_projet_id'] = projet.chef_projet.id
    data['client_nom'] = projet.client.nom
    return JsonResponse(data)










def get_competences_certificats(request):
    if request.method == "GET":
        competences = list(Competence.objects.values('id', 'nom'))
        certificats = list(Certificat.objects.values('id', 'intitule'))
        return JsonResponse({'competences': competences, 'certificats': certificats})
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)



from django.utils import timezone
def voir_certificats(request, personnel_id):
    agent = get_object_or_404(Personnel, id=personnel_id)
    certificats = agent.certificats.all()
    aujourd_hui = timezone.now().date()
    return render(request, 'certificats_details.html', {'agent': agent, 'certificats': certificats , 'aujourd_hui': aujourd_hui})




def parse_date(date_str):
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return None
    
@csrf_exempt
@transaction.atomic
def enregistrer_agent(request):
    if request.method == 'POST':
        try:
            print("➡️ Requête reçue en POST")

            # 1. Récupération des champs standards
            nom = request.POST.get('nom')
            prenoms = request.POST.get('prenoms')
            email = request.POST.get('email')
            telephone = request.POST.get('telephone')
            nationalite = request.POST.get('nationalite')
            statut = request.POST.get('statut')
            residence = request.POST.get('residence')
            pays_affectation = request.POST.get('pays_affectation')
            libelle_competence = request.POST.get('competence')
           

            # Création du personnel
            personnel = Personnel.objects.create(
                nom=nom,
                prenoms=prenoms,
                email=email,
                telephone=telephone,
                nationalite=nationalite,
                statut=statut,
                residence=residence
            )

            # 2. Gestion du pays d'affectation → créer si inexistant
            if pays_affectation:
                pays, _ = PaysAffectation.objects.get_or_create(nom_pays=pays_affectation)
                # Date d’expatriation aujourd’hui par défaut (ou à adapter)
                Expatriation.objects.create(
                    personnel=personnel,
                    pays=pays,
                    date_expatriation=parse_date(request.POST.get('date_expatriation')) or timezone.now().date()
                )

            libelle_competence = request.POST.get('competence')
            print(f"Valeur de libelle_competence reçue : '{libelle_competence}'") # <<< Ajoutez ceci

            if libelle_competence:
                
                    competence, _ = Competence.objects.get_or_create(libelle=libelle_competence)
                    print(f"Competence trouvée/créée : {competence.libelle}, Créée: {_}")
                    Posseder.objects.create(personnel=personnel, competence=competence)
                    print("Liaison Posseder créée.")
            # 4. Gestion des certificats dynamiques
            certificats = [key.split('[')[1].split(']')[0] for key in request.POST if key.startswith('certificats[')]
            certificats = list(set(certificats))  # Supprimer les doublons

            for cert_id in certificats:
                prefix = f'certificats[{cert_id}]'
                libelle = request.POST.get(f'{prefix}[libelle]')
                type_cert = request.POST.get(f'{prefix}[type]')
                obtention = parse_date(request.POST.get(f'{prefix}[obtention]'))
                validite = parse_date(request.POST.get(f'{prefix}[validite]'))
                organisme = request.POST.get(f'{prefix}[organisme]')
                fichier = request.FILES.get(f'{prefix}[fichier]')
                # Vérification des champs requis
                print(f"Traitement du certificat {cert_id}: {libelle}, {type_cert}, {obtention}, {validite}, {organisme}")
                if all([libelle, type_cert, obtention, validite, organisme]):
                 Certificat.objects.create(
                    personnel=personnel,
                    libelle=libelle,
                    type=type_cert,
                    date_obtention=obtention,
                    validite=validite,
                    organisme=organisme,
                    fichier_pdf=fichier
                )

            return JsonResponse({'success': True, 'message': 'Agent enregistré avec succès.'})

        except Exception as e:
            print("❌ Erreur :", str(e))
            return JsonResponse({'success': False, 'errors': {'exception': str(e)}}, status=400)

    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)






def modifier_agent(request, agent_id):
    agent = get_object_or_404(Personnel, pk=agent_id)

    if request.method == 'POST':
        form = ModifierPersonnelForm(request.POST, request.FILES, agent=agent, instance=agent)
        if form.is_valid():
            form.save()
            return redirect('tabpersonels')
    else:
        form = ModifierPersonnelForm(agent=agent, instance=agent)

    return render(request, 'modifier_agent.html', {
        'form': form,
        'agent': agent
    })
    
    


@require_POST
@csrf_exempt
def supprimer_agent(request, agent_id):
    try:
        agent = get_object_or_404(Personnel, id=agent_id)
        agent.delete()
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)







def creer_equipe(request):
    competence = request.GET.get('competence')  # ?competence=RAN par ex.

    if request.method == 'POST':
        equipe_form = EquipeForm(request.POST)
        formset = MembreFormSet(
            request.POST,
            queryset=Membre.objects.none(),
            form_kwargs={'competence': competence}  
        )

        if equipe_form.is_valid() and formset.is_valid():
            membres_valides = []
            for form in formset:
                if form.cleaned_data.get('personnel') and form.cleaned_data.get('role'):
                    membres_valides.append((
                        form.cleaned_data['personnel'].id,
                        form.cleaned_data['role']
                    ))

            if len(membres_valides) < 3:
                messages.warning(request, "Une équipe doit contenir au moins 3 membres.")
            elif len(set(m[0] for m in membres_valides)) != len(membres_valides):
                messages.warning(request, "Un même membre ne peut pas apparaître plusieurs fois.")
            else:
                signature_nouvelle = set(membres_valides)
                for equipe in Equipe.objects.all():
                    signature_existante = set(
                        (m.personnel.id, m.role)
                        for m in equipe.membres.all()
                    )
                    if signature_existante == signature_nouvelle:
                        messages.warning(request, f"L'équipe « {equipe.nom} » avec ces membres et rôles existe déjà.")
                        break
                else:
                    equipe = equipe_form.save()
                    for form in formset:
                        cd = form.cleaned_data
                        if cd.get('personnel') and cd.get('role'):
                            membre = form.save(commit=False)
                            membre.equipe = equipe
                            membre.save()
                    messages.success(request, "Équipe créée avec succès.")
                    return redirect('liste_equipes')

    else:
        equipe_form = EquipeForm()
        formset = MembreFormSet(
            queryset=Membre.objects.none(),
            form_kwargs={'competence': competence}  
        )

    return render(request, 'equipes.html', {
        'equipe_form': equipe_form,
        'formset': formset,
        'competence': competence,
    })









def modifier_equipe(request, equipe_id):
    # Récupère l'équipe à modifier ou renvoie une erreur 404 si elle n'existe pas
    equipe = get_object_or_404(Equipe, id=equipe_id)

    if request.method == 'POST':
        # Si le formulaire est soumis (méthode POST)
        form = EquipeForm(request.POST, instance=equipe)
        if form.is_valid():
            form.save()
            # Redirigez vers une page de succès ou la liste des équipes
            return redirect('liste_equipes') # Assurez-vous que cette URL existe
    else:
        # Si c'est une requête GET (première affichage du formulaire)
        form = EquipeForm(instance=equipe) # Pré-remplit le formulaire avec les données de l'équipe

    return render(request, 'modifier_equipe.html', {'form': form, 'equipe': equipe})


def supprimer_equipe(request, equipe_id):
    # Récupère l'équipe à supprimer ou renvoie une erreur 404
    equipe = get_object_or_404(Equipe, id=equipe_id)

    if request.method == 'POST':
        # La suppression est généralement confirmée via un POST
        equipe.delete()
        # Redirigez vers une page de succès ou la liste des équipes
        return redirect('liste_equipes') # Assurez-vous que cette URL existe
    
    else:
        
        return redirect('liste_equipes')


def liste_equipes(request):
    equipes = Equipe.objects.prefetch_related('membres__personnel')

    return render(request, 'liste_equipes.html', {'equipes': equipes})


from django.views.decorators.http import require_GET

@require_GET
def equipes_disponibles(request):
    projet_id = request.GET.get('projet_id')
    if projet_id:
        equipe_ids = AffectationProjet.objects.filter(projet_id=projet_id).values_list('equipe_id', flat=True)
        equipes = Equipe.objects.exclude(id__in=equipe_ids).values('id', 'nom')
        return JsonResponse(list(equipes), safe=False)
    return JsonResponse([], safe=False)

@require_GET
def projets_disponibles(request):
    equipe_id = request.GET.get('equipe_id')
    if equipe_id:
        projet_ids = AffectationProjet.objects.filter(equipe_id=equipe_id).values_list('projet_id', flat=True)
        projets = Projet.objects.filter(statut='en_cours').exclude(id__in=projet_ids).values('id', 'nom')
        return JsonResponse(list(projets), safe=False)
    return JsonResponse([], safe=False)






def affecter_equipe(request):
    
    if request.method == 'POST':
        form = AffectationProjetForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Équipe affectée au projet avec succès.")
            return redirect('affecter_equipe')  # Redirige pour éviter la double soumission
        else:
            # Si le formulaire n'est pas valide, messages.error peut afficher les erreurs de validation
            # Le formulaire contiendra automatiquement les erreurs non liées aux champs (comme celle de clean())
            # et les erreurs spécifiques aux champs.
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = AffectationProjetForm()

    affectations = AffectationProjet.objects.select_related('projet', 'equipe').order_by('-date_affectation')

    return render(request, 'affectation_projet.html', {
        'form': form,
        'affectations': affectations

    })



def supprimer_affectation(request, affectation_id):
    affectation = get_object_or_404(AffectationProjet, id=affectation_id)

    # Vérifie si le projet est encore en cours
    if affectation.projet.statut != 'en_cours':
        messages.error(request, "Impossible de supprimer une affectation pour un projet terminé.")
        return redirect('affecter_equipe')

    if request.method == 'POST':
        affectation.delete()
        messages.success(request, "Affectation supprimée avec succès.")
        return redirect('affecter_equipe')



from datetime import date
from django.shortcuts import render, get_object_or_404, redirect
from .models import Projet, Equipe, Membre, Effectuer, Activite
from .forms import AffecterMembreTacheForm

def affecter_membres_tache(request, projet_id, equipe_id):
    projet = get_object_or_404(Projet, id=projet_id)
    equipe = get_object_or_404(Equipe, id=equipe_id)

    membres_de_l_equipe = Membre.objects.filter(equipe=equipe)

    if request.method == 'POST':
        form = AffecterMembreTacheForm(request.POST, projet=projet)
        if form.is_valid():
            activite = form.cleaned_data['activite']
            membres = form.cleaned_data['membres']

            for membre in membres:
                Effectuer.objects.update_or_create(
                    membre=membre,
                    activite=activite,
                    defaults={'date_affecter': date.today()}  # ✅ Date automatique
                )
            return redirect('affecter_membres_tache', projet_id=projet.id, equipe_id=equipe.id)
    else:
        form = AffecterMembreTacheForm(projet=projet)

    affectations = Effectuer.objects.filter(activite__projet=projet).select_related('membre', 'activite')

    return render(request, 'affectation_tache.html', {
        'form': form,
        'projet': projet,
        'equipe': equipe,
        'affectations': affectations,
        'membres_de_l_equipe': membres_de_l_equipe
    })



# def allouer_taches(request, projet_id, equipe_id):
#     projet = get_object_or_404(Projet, id=projet_id)
#     equipe = get_object_or_404(Equipe, id=equipe_id)

#     # Récupérer toutes les tâches associées à ce projet
#     taches_du_projet = Activite.objects.filter(projet=projet)

#     # Récupérer les membres de l'équipe
#     membres_equipe = Membre.objects.filter(equipe=equipe)

#     if not membres_equipe:
#         messages.warning(request, f"L'équipe '{equipe.nom}' n'a aucun membre pour allouer les tâches.")
#         return redirect('nom_de_votre_page_affectation') # Redirigez vers la page d'où vous venez

#     if not taches_du_projet:
#         messages.info(request, f"Le projet '{projet.nom}' n'a pas de tâches non allouées.")
#         return redirect('nom_de_votre_page_affectation')

#     # Logique d'allocation des tâches
#     # C'est la partie la plus importante et elle dépendra de vos besoins :
#     # - Répartition égale ?
#     # - En fonction des compétences des membres ?
#     # - Manuelle par une interface ?

#     # Exemple simple : Répartir les tâches une par une aux membres de l'équipe
#     # Vous devrez adapter cette logique pour la rendre plus intelligente
#     index_membre = 0
#     taches_allouees_count = 0
#     for tache in taches_du_projet:
#         if membres_equipe: # S'assurer qu'il y a des membres à qui allouer
#             membre_a_allouer = membres_equipe[index_membre]
#             tache.assigne_a = membre_a_allouer # Supposons un champ ForeignKey 'assigne_a' dans votre modèle Tache
#             tache.est_allouee = True # Marquez la tâche comme allouée
#             tache.save()
#             taches_allouees_count += 1
#             index_membre = (index_membre + 1) % len(membres_equipe) # Passer au membre suivant, ou revenir au début

#     if taches_allouees_count > 0:
#         messages.success(request, f"{taches_allouees_count} tâches du projet '{projet.nom}' ont été allouées aux membres de l'équipe '{equipe.nom}'.")
#     else:
#         messages.info(request, f"Aucune nouvelle tâche n'a été allouée pour le projet '{projet.nom}' et l'équipe '{equipe.nom}'.")

#     return redirect('nom_de_votre_page_affectation') # Assurez-vous de remplacer par le nom de l'URL de votre page actuelle



from django.shortcuts import render, get_object_or_404, redirect
from .models import Projet, Activite
from .forms import ActiviteForm, LivrableInlineFormSet
from django.forms import formset_factory

def ajouter_taches(request, projet_id):
    projet = get_object_or_404(Projet, id=projet_id)

    class BoundActiviteForm(ActiviteForm):
        def __init__(self, *args, **kwargs):
            kwargs.setdefault('projet', projet)
            super().__init__(*args, **kwargs)

    TacheFormSet = formset_factory(BoundActiviteForm, extra=1, can_delete=True)

    if request.method == "POST":
        formset = TacheFormSet(request.POST, prefix="taches")

        if formset.is_valid():
            for i, form in enumerate(formset):
                if form.cleaned_data.get("DELETE"):
                    continue
                if form.has_changed():
                    activite = form.save(commit=False)
                    
                    # ✅ On affecte le projet avant de sauvegarder ou d'utiliser activite.projet
                    activite.projet = projet
                    activite.statut = "planifiée"
                    activite.save()  # Maintenant, activite a bien un ID et un projet
                    
                    # ✅ Gestion des livrables
                    livrable_formset = LivrableInlineFormSet(
                        request.POST,
                        instance=activite,
                        prefix=f"livrables-{i}"
                    )

                    if livrable_formset.is_valid():
                        livrable_formset.save()
                    else:
                        print("Erreur dans les livrables :", livrable_formset.errors)

            return redirect("tabprojet")
    else:
        formset = TacheFormSet(prefix="taches")

    context = {
        "projet": projet,
        "formset": formset,
    }
    return render(request, "ajouter_taches.html", context)



import openpyxl
from django.http import HttpResponse
from .models import Projet, Activite, Livrable


def exporter_projet_excel(request, projet_id):
    projet = Projet.objects.get(id=projet_id)

    # Créer un classeur Excel
    workbook = openpyxl.Workbook()

    # Feuille principale du projet
    sheet = workbook.active
    sheet.title = "Projet"
    sheet.append([
        "Nom du projet", "Date de début", "Date de fin", "Statut"
    ])
    sheet.append([
        projet.nom,
        str(projet.date_debut),
        str(projet.date_fin),
        projet.statut
    ])

    # Ajouter les tâches dans une autre feuille
    sheet_taches = workbook.create_sheet("Tâches")
    sheet_taches.append([
        "ID Tâche", "Nom", "Description", "Date Début", "Date Fin", "Statut", "Tâche Précédente"
    ])

    for activite in Activite.objects.filter(projet=projet):
        sheet_taches.append([
            activite.id,
            activite.nom,
            activite.description,
            str(activite.date_debut),
            str(activite.date_fin),
            activite.statut,
            activite.tache_precedente.nom if activite.tache_precedente else ""
        ])

        # Ajouter les livrables liés à la tâche dans leur propre feuille
        sheet_livrables = workbook.create_sheet(f"Liv.{activite.id}")
        sheet_livrables.append(["ID Livrable", "Nom"])

        for livrable in Livrable.objects.filter(activite=activite):
            sheet_livrables.append([
                livrable.id,
                livrable.nom_livrable,
              
            ])

    # Réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Projet_{projet.nom}.xlsx'

    workbook.save(response)
    return response



def mobiliser_equipes(request, projet_id):
    projet = get_object_or_404(Projet, id=projet_id)
    equipes = projet.equipe_affectee.all()

    return render(request, 'mobilisation.html', {
        'projet': projet,
        'equipes': equipes,
    })





from django.shortcuts import get_object_or_404, render, redirect
from django.db import transaction
from django.contrib import messages
from .models import Projet, Activite, Equipe, AffectationProjet, Realiser, Mobilisation
from .forms import MobilisationForm

def creer_mobilisation(request, projet_id):
    projet = get_object_or_404(Projet, id=projet_id)
    equipes_pre_affectees = Equipe.objects.filter(affectations_projet__projet=projet)

    # Liste des activités du projet
    activites = projet.activites.all()

    if request.method == 'POST':
        activite_id = request.POST.get('activite')
        mobilisation_form = MobilisationForm(request.POST)
        equipe_ids = request.POST.getlist('equipes_selectionnees')

        # Validation de la sélection de l'activité
        if not activite_id:
            messages.error(request, "Veuillez sélectionner une activité.")
            return render(request, 'organiser_mobilisation.html', {
                'projet': projet,
                'equipes_pre_affectees': equipes_pre_affectees,
                'activites': activites,
                'mobilisation_form': mobilisation_form,
            })

        activite = get_object_or_404(Activite, id=activite_id, projet=projet)

        if not equipe_ids:
            messages.error(request, "Veuillez sélectionner au moins une équipe.")
            return render(request, 'organiser_mobilisation.html', {
                'projet': projet,
                'equipes_pre_affectees': equipes_pre_affectees,
                'activites': activites,
                'mobilisation_form': mobilisation_form,
            })

        if mobilisation_form.is_valid():
            try:
                with transaction.atomic():
                    # Enregistrer la mobilisation
                    mobilisation = mobilisation_form.save(commit=False)
                    mobilisation.activite = activite
                    mobilisation.chef_projet = request.user.chef_projet
                    mobilisation.save()

                    # Associer les équipes à l'activité via Realiser
                    for equipe_id in equipe_ids:
                        equipe = get_object_or_404(Equipe, id=equipe_id)
                        Realiser.objects.create(
                            equipe=equipe,
                            activite=activite
                        )

                    messages.success(request, "Mobilisation enregistrée avec succès.")
                    return redirect('resume_mobilisation', mobilisation_id=mobilisation.id)  

            except Exception as e:
                messages.error(request, f"Une erreur est survenue : {e}")
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")

    else:
        mobilisation_form = MobilisationForm()

    return render(request, 'organiser_mobilisation.html', {
        'projet': projet,
        'equipes_pre_affectees': equipes_pre_affectees,
        'activites': activites,
        'mobilisation_form': mobilisation_form,
    })






def organiser_mobilisation(request):
    
    # Récupérer les projets "en cours" ayant au moins une affectation liée
    projets = Projet.objects.filter(
        statut='en_cours',
        equipe_affectee__isnull=False  # relation inversée de AffectationProjet
    ).distinct()  # éviter les doublons si plusieurs équipes
    print("Projet:", projets)

    #  Préparation des données pour le template
    data = []
    for projet in projets:
        # On récupère les affectations pour ce projet
        affectations = AffectationProjet.objects.filter(projet=projet).select_related('equipe')

        # On extrait les équipes à partir des affectations
        equipes = [a.equipe for a in affectations]
        print("Projet:", projet.nom, "Équipes:", [e.nom for e in equipes])
        # On construit une entrée dans la liste finale
        data.append({
            'projet': projet,
            'equipes': equipes
        })

    return render(request, 'mobilisation.html', {'projets_data': data})




def resume_mobilisation(request, mobilisation_id):
    mobilisation = get_object_or_404(Mobilisation, id=mobilisation_id)
    activite = mobilisation.activite
    equipes = mobilisation.activite.equipes_realisation.all()

    return render(request, 'resume_mobilisation.html', {
        'mobilisation': mobilisation,
        'activite': activite,
        'equipes': equipes,
    })




from django.db.models import Prefetch
def liste_mobilisations(request):
    #mobilisations = Mobilisation.objects.select_related('activite', 'chef_projet').all().order_by('-date_debut')
    mobilisations = Mobilisation.objects.select_related('activite', 'chef_projet').prefetch_related(Prefetch('activite__equipes_realisation', queryset=Realiser.objects.select_related('equipe'), to_attr='prefetched_equipes_realisation'  )  ).order_by('-date_debut')# Précharger les équipes via Realiser et Activite si besoin
    context = {
        'mobilisations': mobilisations,
    }
    return render(request, 'liste_mobilisations.html', context)





def liste_activites(request):
    activites = Activite.objects.select_related('projet').prefetch_related('equipes_realisation', 'livrables').filter(projet__statut='en_cours').order_by('created_at')

    return render(request, 'activites.html', {
        'activites': activites,
        'statut_choices': STATUT_ACTIVITIES_CHOICES,
    })




def changer_statut_activite(request, activite_id):
    activite = get_object_or_404(Activite, id=activite_id)
    nouveau_statut = request.POST.get('statut')

    if nouveau_statut in dict(STATUT_ACTIVITIES_CHOICES):
        activite.statut = nouveau_statut
        activite.save()
        return JsonResponse({'success': True, 'message': 'Statut mis à jour.'})
    return JsonResponse({'success': False, 'message': 'Statut non valide.'}, status=400)










from django.db.models import Q
def search_results(request):
    query = request.GET.get('q') # Récupère le terme de recherche
    
    # Initialise un dictionnaire pour stocker les résultats de chaque catégorie
    results = {
        'projets': [],
        'personnel': [],
        'equipes': [],
    }

    if query:
        # --- Recherche dans le modèle Projet ---
        projets_filters = Q() # Crée un objet Q vide pour construire les conditions

        # Champs textuels : Utilisez __icontains
        projets_filters |= Q(nom__icontains=query) 

        # Si 'client' est un CharField (un simple champ texte dans le modèle Projet)
        # projets_filters |= Q(client__icontains=query) 
        
        # SI 'client' est une ForeignKey vers un modèle 'Client' qui a un champ 'nom' (ou 'entreprise', 'raison_sociale'...)
        # Alors, traversez la relation :
        projets_filters |= Q(client__nom__icontains=query) 
        # projets_filters |= Q(client__entreprise__icontains=query) # Exemple si le champ est 'entreprise'

        # Champ de date (date_fin) : 
        # Vous devez gérer les dates différemment.
        # Par exemple, si l'utilisateur entre une année :
        try:
            year = int(query)
            projets_filters |= Q(date_fin__year=year) # Recherche par année exacte
        except ValueError:
            # Si 'query' n'est pas un nombre, ignore la recherche par année.
            pass
        # Ou si vous voulez permettre une recherche par mois/jour, c'est plus complexe et dépend
        # du format de la requête de l'utilisateur.

        projets_results = Projet.objects.filter(projets_filters).distinct()


        # --- Recherche dans le modèle Personnel ---
        personnel_filters = Q()

        # Champs textuels et email : Utilisez __icontains
        personnel_filters |= Q(nom__icontains=query) 
        personnel_filters |= Q(prenoms__icontains=query)
        personnel_filters |= Q(email__icontains=query)

        # Champ téléphone (chiffres, souvent stocké en CharField pour garder les + et les espaces)
        # Si 'telephone' est un CharField :
        personnel_filters |= Q(telephone__icontains=query)
        # Si 'telephone' est un IntegerField, vous devrez le convertir en chaîne pour la recherche,
        # mais c'est moins commun et moins performant que de le stocker en CharField.

        # Champ 'statut' :
        # Si 'statut' est un CharField (ex: 'Actif', 'En congé')
        personnel_filters |= Q(statut__icontains=query) 
        
        # SI 'statut' est une ForeignKey vers un modèle 'Statut' (ex: avec un champ 'libelle' ou 'nom')
        # ALORS traversez la relation :
        # personnel_filters |= Q(statut__libelle__icontains=query) 
        # personnel_filters |= Q(statut__nom__icontains=query)

        personnel_results = Personnel.objects.filter(personnel_filters).distinct()


        # --- Recherche dans le modèle Equipe ---
        equipes_filters = Q()

        # Champ textuel : Utilisez __icontains
        equipes_filters |= Q(nom__icontains=query)

        # Champ de date (date_creation) : NE PAS utiliser __icontains.
        # Même logique que pour 'date_fin' dans Projet.
        try:
            year = int(query)
            equipes_filters |= Q(date_creation__year=year) # Recherche par année exacte
        except ValueError:
            pass

        equipes_results = Equipe.objects.filter(equipes_filters).distinct()


        # Assignation des résultats
        results['projets'] = projets_results
        results['personnel'] = personnel_results
        results['equipes'] = equipes_results

    # Rend le template search_results.html avec la requête et les résultats
    return render(request, 'recherche_sur_lesite.html', {'query': query, 'results': results})




def logout_view(request):
    return redirect('login')
    
